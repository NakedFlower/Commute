"""
Slack 출퇴근 관리 시스템
슬래시 커맨드(/출근, /외근, /퇴근)를 통해 Excel 파일에 출퇴근 시간을 자동 기록
"""

from dotenv import load_dotenv
load_dotenv()

import os
import hmac
import hashlib
import time
from datetime import datetime
from typing import Optional
from threading import Lock
from pathlib import Path

from fastapi import FastAPI, Request, HTTPException, Form
from fastapi.responses import JSONResponse
import pytz
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from slack_sdk import WebClient
from slack_sdk.errors import SlackApiError
from urllib.parse import parse_qs

# ============================================================================
# 환경 변수 설정
# ============================================================================
SLACK_SIGNING_SECRET = os.getenv("SLACK_SIGNING_SECRET", "")
SLACK_BOT_TOKEN = os.getenv("SLACK_BOT_TOKEN", "")

if not SLACK_SIGNING_SECRET:
    print("⚠️ WARNING: SLACK_SIGNING_SECRET이 설정되지 않았습니다.")
if not SLACK_BOT_TOKEN:
    print("⚠️ WARNING: SLACK_BOT_TOKEN이 설정되지 않았습니다.")

# ============================================================================
# 전역 설정
# ============================================================================
EXCEL_FILE = "attendance.xlsx"
SHEET_NAME = "Attendance"
KST = pytz.timezone("Asia/Seoul")

# Excel 파일 접근용 Lock (동시성 제어)
file_lock = Lock()

# Slack 클라이언트
slack_client = WebClient(token=SLACK_BOT_TOKEN) if SLACK_BOT_TOKEN else None

# FastAPI 앱 초기화
app = FastAPI(title="Slack 출퇴근 관리 시스템")


# ============================================================================
# Slack 서명 검증
# ============================================================================
def verify_slack_signature(request_body: bytes, timestamp: str, signature: str) -> bool:
    """
    Slack 요청의 서명을 검증하여 위조 요청 방지
    
    Args:
        request_body: 요청 본문
        timestamp: Slack이 보낸 타임스탬프
        signature: Slack이 보낸 서명
        
    Returns:
        bool: 서명이 유효하면 True
    """
    if not SLACK_SIGNING_SECRET:
        # 개발 환경에서 시크릿이 없으면 검증 스킵
        print("⚠️ 서명 검증 스킵 (SLACK_SIGNING_SECRET 없음)")
        return True
    
    # 타임스탬프가 5분 이상 오래된 요청은 거부 (재생 공격 방지)
    if abs(time.time() - int(timestamp)) > 60 * 5:
        return False
    
    # 서명 생성
    sig_basestring = f"v0:{timestamp}:{request_body.decode('utf-8')}"
    my_signature = 'v0=' + hmac.new(
        SLACK_SIGNING_SECRET.encode(),
        sig_basestring.encode(),
        hashlib.sha256
    ).hexdigest()
    
    # 서명 비교 (timing attack 방지)
    return hmac.compare_digest(my_signature, signature)


# ============================================================================
# Excel 파일 관리
# ============================================================================
def init_excel_file():
    """
    Excel 파일이 없으면 생성하고 헤더 설정
    """
    if Path(EXCEL_FILE).exists():
        return
    
    with file_lock:
        # 이중 체크 (lock 획득 후 다시 확인)
        if Path(EXCEL_FILE).exists():
            return
        
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        
        # 헤더 설정
        headers = ["날짜", "Slack User ID", "이름", "출근 시간", "외근 시간", "퇴근 시간"]
        ws.append(headers)
        
        # 헤더 스타일링
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 컬럼 너비 조정
        ws.column_dimensions['A'].width = 12  # 날짜
        ws.column_dimensions['B'].width = 15  # User ID
        ws.column_dimensions['C'].width = 15  # 이름
        ws.column_dimensions['D'].width = 12  # 출근
        ws.column_dimensions['E'].width = 12  # 외근
        ws.column_dimensions['F'].width = 12  # 퇴근
        
        wb.save(EXCEL_FILE)
        print(f"✅ Excel 파일 생성 완료: {EXCEL_FILE}")


def get_user_display_name(user_id: str) -> str:
    """
    Slack User ID로 실제 이름 조회
    
    Args:
        user_id: Slack User ID
        
    Returns:
        str: 사용자 이름 (조회 실패 시 User ID 반환)
    """
    if not slack_client:
        return user_id
    
    try:
        response = slack_client.users_info(user=user_id)
        user_info = response["user"]
        # real_name > display_name > name 순서로 조회
        return (
            user_info.get("real_name") or 
            user_info.get("profile", {}).get("display_name") or 
            user_info.get("name", user_id)
        )
    except SlackApiError as e:
        print(f"⚠️ 사용자 정보 조회 실패: {e}")
        return user_id


def record_attendance(user_id: str, command_type: str) -> str:
    """
    출퇴근 시간을 Excel 파일에 기록
    
    Args:
        user_id: Slack User ID
        command_type: "출근", "외근", "퇴근" 중 하나
        
    Returns:
        str: 기록된 시간 (HH:MM 형식)
    """
    # 현재 시간 (KST)
    now = datetime.now(KST)
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H:%M")
    
    # 사용자 이름 조회
    user_name = get_user_display_name(user_id)
    
    # 컬럼 매핑
    column_map = {
        "출근": 3,  # D열 (0-indexed: 3)
        "외근": 4,  # E열
        "퇴근": 5   # F열
    }
    
    if command_type not in column_map:
        raise ValueError(f"알 수 없는 커맨드 타입: {command_type}")
    
    target_col = column_map[command_type]
    
    with file_lock:
        # Excel 파일 로드
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]
        
        # 기존 레코드 찾기 (같은 날짜 + 같은 사용자)
        target_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if row[0].value == date_str and row[1].value == user_id:
                target_row = row_idx
                break
        
        # 기존 레코드가 없으면 새 행 추가
        if target_row is None:
            target_row = ws.max_row + 1
            ws.cell(row=target_row, column=1, value=date_str)  # 날짜
            ws.cell(row=target_row, column=2, value=user_id)   # User ID
            ws.cell(row=target_row, column=3, value=user_name) # 이름
        
        # 해당 시간 컬럼에 기록
        ws.cell(row=target_row, column=target_col + 1, value=time_str)
        
        # 중앙 정렬
        for col in range(1, 7):
            ws.cell(row=target_row, column=col).alignment = Alignment(
                horizontal="center", 
                vertical="center"
            )
        
        # 저장
        wb.save(EXCEL_FILE)
    
    return time_str


# ============================================================================
# FastAPI 엔드포인트
# ============================================================================
@app.on_event("startup")
async def startup_event():
    """
    서버 시작 시 Excel 파일 초기화
    """
    init_excel_file()
    print("🚀 서버 시작 완료")


@app.get("/")
async def root():
    """
    헬스체크 엔드포인트
    """
    return {"status": "ok", "message": "Slack 출퇴근 관리 시스템이 정상 작동 중입니다."}

@app.post("/slack/commands")
async def slack_commands(request: Request):
    # 1️⃣ RAW BODY 먼저 읽기 (딱 한 번)
    body = await request.body()

    # 2️⃣ Slack 서명 검증
    timestamp = request.headers.get("X-Slack-Request-Timestamp", "")
    signature = request.headers.get("X-Slack-Signature", "")

    if not verify_slack_signature(body, timestamp, signature):
        raise HTTPException(status_code=401, detail="Invalid signature")

    # 3️⃣ body를 Slack form 데이터로 직접 파싱
    form = parse_qs(body.decode())

    command = form.get("command", [""])[0]
    text = form.get("text", [""])[0]
    user_id = form.get("user_id", [""])[0]
    user_name = form.get("user_name", [""])[0]

    # 4️⃣ 커맨드 처리
    command_type = command.lstrip("/")

    if command_type not in ["출근", "외근", "퇴근"]:
        return JSONResponse(content={
            "response_type": "ephemeral",
            "text": f"❌ 지원하지 않는 커맨드입니다: {command}"
        })

    try:
        recorded_time = record_attendance(user_id, command_type)

        emoji_map = {
            "출근": "🏢",
            "외근": "🚗",
            "퇴근": "🏠"
        }

        return JSONResponse(content={
            "response_type": "ephemeral",
            "text": f"{emoji_map[command_type]} **{command_type}** 시간이 **{recorded_time}**로 기록되었습니다."
        })

    except Exception as e:
        print("❌ 오류:", e)
        return JSONResponse(content={
            "response_type": "ephemeral",
            "text": f"❌ 오류 발생: {str(e)}"
        })


# ============================================================================
# 서버 실행 (개발 환경)
# ============================================================================
if __name__ == "__main__":
    import uvicorn
    
    print("=" * 60)
    print("🚀 Slack 출퇴근 관리 시스템 시작")
    print("=" * 60)
    print(f"📁 Excel 파일: {EXCEL_FILE}")
    print(f"🌐 서버 주소: http://localhost:8000")
    print(f"🔗 엔드포인트: http://localhost:8000/slack/commands")
    print("=" * 60)
    
    uvicorn.run(app, host="0.0.0.0", port=8000)
